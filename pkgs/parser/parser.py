#!/usr/bin/env python3
"""
Парсер расписания из Excel-файлов на Google Диске.
Читает файлы .xlsx напрямую в память, анализирует обводку ячеек.
Результат сохраняется в parsed_schedule.json и отправляется на backend (Cloudflare Worker).
"""

import os
import io
import re
import json
import hashlib
import time
import uuid
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
from itertools import groupby

import requests

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# 1.  КОНФИГУРАЦИЯ И ПЕРЕМЕННЫЕ ОКРУЖЕНИЯ
# ============================================================

GOOGLE_SERVICE_ACCOUNT_JSON = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
DRIVE_FOLDER_ID = os.environ.get("DRIVE_FOLDER_ID")
BACKEND_URL = os.environ.get("BACKEND_URL")
BACKEND_AUTH_TOKEN = os.environ.get("BACKEND_AUTH_TOKEN")

if not GOOGLE_SERVICE_ACCOUNT_JSON:
    raise ValueError("Не задана переменная GOOGLE_SERVICE_ACCOUNT_JSON")
if not DRIVE_FOLDER_ID:
    raise ValueError("Не задана переменная DRIVE_FOLDER_ID")

if not BACKEND_URL:
    print("⚠️ BACKEND_URL не задан — данные будут только сохранены локально.")
if not BACKEND_AUTH_TOKEN:
    print("⚠️ BACKEND_AUTH_TOKEN не задан — отправка на backend будет пропущена.")

# Стандартная высота одной пары в строках таблицы: title / препод / препод / аудитория.
STANDARD_SLOT_HEIGHT = 4
# Сколько лишних строк сверх стандарта допускаем на уровне разметки дня, прежде чем
# извлечение пары само разберётся, откуда резать (см. extract_class). Один запас нужен
# для случая "в таблице по ошибке добавлена лишняя пустая строка". Без этого потолка
# последний слот последнего дня (суббота) мог бы "утекать" в пустые строки листа ниже
# таблицы (next_day_min_row = ws.max_row + 1 для последнего дня).
MAX_SLOT_ROW_SLACK = 1


# ============================================================
# 2.  АВТОРИЗАЦИЯ В GOOGLE DRIVE
# ============================================================


def get_drive_service():
    try:
        creds_dict = json.loads(GOOGLE_SERVICE_ACCOUNT_JSON)
        creds = service_account.Credentials.from_service_account_info(
            creds_dict, scopes=["https://www.googleapis.com/auth/drive.readonly"]
        )
        return build("drive", "v3", credentials=creds)
    except Exception as e:
        raise RuntimeError(f"Ошибка авторизации в Google Drive: {e}")


# ============================================================
# 3.  ПОЛУЧЕНИЕ СПИСКА ФАЙЛОВ В ПАПКЕ
# ============================================================


def list_files_in_folder(service, folder_id: str) -> List[Dict]:
    query = f"'{folder_id}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
    results = (
        service.files()
        .list(q=query, fields="files(id, name, mimeType)", pageSize=50)
        .execute()
    )
    return results.get("files", [])


def extract_course_number(filename: str) -> Optional[int]:
    match = re.search(r"(\d+)\s*курс", filename, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None


# ============================================================
# 4.  ЧТЕНИЕ ФАЙЛА В ПАМЯТЬ
# ============================================================


def read_file_as_bytes(service, file_id: str) -> bytes:
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    return fh.getvalue()


# ============================================================
# 5.  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ РАБОТЫ С ЯЧЕЙКАМИ
# ============================================================


def get_merged_range_for_cell(ws: Worksheet, row: int, col: int):
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return merged_range
    return None


def get_cell_value(ws: Worksheet, row: int, col: int) -> Any:
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    merged = get_merged_range_for_cell(ws, row, col)
    if merged:
        return ws.cell(row=merged.min_row, column=merged.min_col).value
    return None


def get_rect_values(
    ws: Worksheet, min_row: int, min_col: int, max_row: int, max_col: int
) -> List[List[Any]]:
    result = []
    for r in range(min_row, max_row + 1):
        row_vals = []
        for c in range(min_col, max_col + 1):
            row_vals.append(get_cell_value(ws, r, c))
        result.append(row_vals)
    return result


def clean_group_number(name: str) -> str:
    cleaned = re.sub(r"\s+", "", str(name))
    cleaned = re.sub(r"группа", "", cleaned, flags=re.IGNORECASE)
    return cleaned


def parse_time_range(label: str) -> Tuple[str, str]:
    cleaned = str(label).strip()
    cleaned = re.sub(r"\s*-\s*", "-", cleaned)
    if "-" in cleaned:
        parts = cleaned.split("-", 1)
        if len(parts) == 2 and parts[0] and parts[1]:
            return parts[0], parts[1]
    return cleaned, ""


# ============================================================
# 6.  ПАРСЕР ДЛЯ 1-ГО И 2-ГО КУРСОВ
# ============================================================


def find_course_range(ws: Worksheet, course_number: int):
    columns_to_check = [1, 2, 3]
    max_search_row = 13

    for col in columns_to_check:
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col == col and merged_range.min_row <= max_search_row:
                cell_value = ws.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                ).value
                if cell_value and f"{course_number} курс" in str(cell_value).lower():
                    height = merged_range.max_row - merged_range.min_row + 1
                    width = merged_range.max_col - merged_range.min_col + 1
                    if 3 <= height <= 6 and 1 <= width <= 4:
                        return merged_range

    for col in columns_to_check:
        for row in range(1, max_search_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and f"{course_number} курс" in str(cell.value).lower():
                for mr in ws.merged_cells.ranges:
                    if cell.coordinate in mr:
                        return mr

                class FakeRange:
                    def __init__(self, min_col, max_col, min_row, max_row):
                        self.min_col = min_col
                        self.max_col = max_col
                        self.min_row = min_row
                        self.max_row = max_row

                return FakeRange(col, col, row, row)
    return None


def parse_course_1_2(file_content: bytes, course_number: int) -> List[Dict]:
    print(f"\n=== Парсинг файла для курса {course_number} ===")
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws: Worksheet = wb.active

    course_range = find_course_range(ws, course_number)
    if not course_range:
        return []

    course_start_row = course_range.min_row
    course_end_col = course_range.max_col

    flows = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row == course_start_row and mr.min_col > course_end_col:
            if (mr.max_col - mr.min_col + 1) >= 2:
                flows.append(mr)
    flows.sort(key=lambda r: r.min_col)

    if not flows:
        return []

    groups_info = []
    for flow_idx, flow in enumerate(flows):
        flow_start_col, flow_end_col = flow.min_col, flow.max_col
        # Строка названий групп для 1-2 курсов сразу под потоком
        group_row = flow.max_row + 1
        col = flow_start_col
        while col <= flow_end_col:
            merged = get_merged_range_for_cell(ws, group_row, col)
            if merged and merged.max_col - merged.min_col + 1 >= 2:
                group_name = ws.cell(row=merged.min_row, column=merged.min_col).value
                if group_name:
                    groups_info.append(
                        {
                            "name": clean_group_number(group_name),
                            "start_col": merged.min_col,
                            "end_col": merged.max_col,
                            "flow": flow_idx,
                            "flow_start_col": flow_start_col,
                            "flow_end_col": flow_end_col,
                        }
                    )
                    col = merged.max_col + 1
                    continue
            col += 1

    if not groups_info:
        return []

    # Подсчет количества групп в каждом потоке
    flow_counts = {}
    for g in groups_info:
        flow_counts[g["flow"]] = flow_counts.get(g["flow"], 0) + 1

    # Определяем последнюю группу в последнем потоке (она не ходит на лекции).
    # Логика идентична parse_course_3_4: только последняя группа последнего потока
    # исключается из лекционного покрытия, для остальных потоков покрытие — весь поток.
    max_flow_idx = max(g["flow"] for g in groups_info)
    last_flow_groups = [g for g in groups_info if g["flow"] == max_flow_idx]
    last_group_no_lectures = None
    if last_flow_groups:
        last_g = max(last_flow_groups, key=lambda g: g["start_col"])
        last_group_no_lectures = (last_g["start_col"], last_g["end_col"])

    # Границы лекций для каждого потока
    flow_lecture_ranges = {}
    for g in groups_info:
        flow_idx = g["flow"]
        if flow_idx not in flow_lecture_ranges:
            if flow_idx == max_flow_idx and last_group_no_lectures:
                groups_in_flow = sorted(
                    [gg for gg in groups_info if gg["flow"] == flow_idx],
                    key=lambda gg: gg["start_col"],
                )
                flow_lecture_ranges[flow_idx] = (
                    (g["flow_start_col"], groups_in_flow[-2]["end_col"])
                    if len(groups_in_flow) >= 2
                    else None
                )
            else:
                flow_lecture_ranges[flow_idx] = (g["flow_start_col"], g["flow_end_col"])

    header_row = course_range.max_row + 1
    day_col, time_col = course_range.min_col, course_range.max_col
    start_data_row = header_row + 1

    days_of_week = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
    day_merged_ranges = []
    for mr in ws.merged_cells.ranges:
        if mr.min_col == day_col and mr.min_row >= start_data_row:
            val = ws.cell(row=mr.min_row, column=mr.min_col).value
            if val and any(day in str(val).lower() for day in days_of_week):
                day_merged_ranges.append(mr)
    day_merged_ranges.sort(key=lambda r: r.min_row)

    days_schedule = []
    for i, day_mr in enumerate(day_merged_ranges):
        day_name = str(ws.cell(row=day_mr.min_row, column=day_mr.min_col).value).strip()
        day_min_row = day_mr.min_row
        day_index = next(
            (idx + 1 for idx, d in enumerate(days_of_week) if d in day_name.lower()),
            i + 1,
        )
        next_day_min_row = (
            day_merged_ranges[i + 1].min_row
            if i + 1 < len(day_merged_ranges)
            else ws.max_row + 1
        )

        day_time_slots = []
        for r in range(day_min_row, next_day_min_row):
            val = ws.cell(row=r, column=time_col).value
            if val is not None and str(val).strip() != "":
                day_time_slots.append({"label": str(val).strip(), "start_row": r})

        # Вычисляем динамическую высоту каждого таймслота, но не даём границе "утечь"
        # дальше разумного предела (стандарт + запас на одну лишнюю строку-ошибку) —
        # точный выбор нужных 4 строк внутри этого диапазона делает extract_class,
        # т.к. там уже известны конкретные колонки конкретной группы.
        for j, slot in enumerate(day_time_slots):
            natural_end = (
                day_time_slots[j + 1]["start_row"] - 1
                if j + 1 < len(day_time_slots)
                else next_day_min_row - 1
            )
            slot["end_row"] = min(
                natural_end,
                slot["start_row"] + STANDARD_SLOT_HEIGHT + MAX_SLOT_ROW_SLACK - 1,
            )

        days_schedule.append(
            {"day_name": day_name, "day_index": day_index, "time_slots": day_time_slots}
        )

    all_parsed = []
    for group in groups_info:
        is_single = flow_counts[group["flow"]] == 1

        is_lecture_participant = True
        lecture_range = flow_lecture_ranges.get(group["flow"])
        if (
            last_group_no_lectures
            and (group["start_col"], group["end_col"]) == last_group_no_lectures
        ):
            is_lecture_participant, lecture_range = False, None
        if not lecture_range:
            is_lecture_participant = False

        for day in days_schedule:
            for slot in day["time_slots"]:
                rect = (
                    slot["start_row"],
                    group["start_col"],
                    slot["end_row"],
                    group["end_col"],
                )
                entry = extract_class(
                    ws,
                    rect,
                    group["name"],
                    course_number,
                    day["day_index"],
                    slot["label"],
                    lecture_coverage_start=lecture_range[0] if lecture_range else None,
                    lecture_coverage_end=lecture_range[1] if lecture_range else None,
                    is_lecture_participant=is_lecture_participant,
                    is_single_group_flow=is_single,
                )
                if entry:
                    all_parsed.extend(entry)

    return all_parsed


# ============================================================
# 7.  ПАРСЕР ДЛЯ 3-ГО И 4-ГО КУРСОВ (УПРОЩЕННЫЙ, КАК ПОДГРУППЫ)
# ============================================================


def parse_course_3_4(file_content: bytes, course_number: int) -> List[Dict]:
    print(f"\n=== Парсинг файла для курса {course_number} ===")
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws: Worksheet = wb.active

    course_range = find_course_range(ws, course_number)
    if not course_range:
        return []

    course_start_row, course_end_col = course_range.min_row, course_range.max_col

    flows = []
    for mr in ws.merged_cells.ranges:
        if (
            mr.min_row == course_start_row
            and mr.min_col > course_end_col
            and (mr.max_col - mr.min_col + 1) >= 2
        ):
            flows.append(mr)
    flows.sort(key=lambda r: r.min_col)

    if not flows:
        return []

    base_groups = []
    for flow_idx, flow in enumerate(flows):
        flow_start_col, flow_end_col = flow.min_col, flow.max_col
        # У 3 и 4 курса строка профилизаций находится под потоком (+1),
        # а названия групп находятся под профилизациями (+2).
        group_row = flow.max_row + 2

        col = flow_start_col
        while col <= flow_end_col:
            merged_group = get_merged_range_for_cell(ws, group_row, col)
            if merged_group and merged_group.max_col - merged_group.min_col + 1 >= 2:
                group_start_col, group_end_col = (
                    merged_group.min_col,
                    merged_group.max_col,
                )
                base_name_raw = ws.cell(
                    row=merged_group.min_row, column=merged_group.min_col
                ).value
                if not base_name_raw:
                    col = merged_group.max_col + 1
                    continue

                clean_base = clean_group_number(base_name_raw)
                base_groups.append(
                    {
                        "base_name": clean_base,
                        "start_col": group_start_col,
                        "end_col": group_end_col,
                        "flow_idx": flow_idx,
                        "flow_start_col": flow_start_col,
                        "flow_end_col": flow_end_col,
                    }
                )
                col = merged_group.max_col + 1
            else:
                col += 1

    if not base_groups:
        return []

    # Подсчет количества групп в каждом потоке
    flow_counts = {}
    for bg in base_groups:
        flow_counts[bg["flow_idx"]] = flow_counts.get(bg["flow_idx"], 0) + 1

    # Определяем последнюю группу в последнем потоке (она не ходит на лекции)
    max_flow_idx = max(bg["flow_idx"] for bg in base_groups)
    last_flow_groups = [bg for bg in base_groups if bg["flow_idx"] == max_flow_idx]
    last_group_no_lectures = (
        (
            max(last_flow_groups, key=lambda bg: bg["start_col"])["start_col"],
            max(last_flow_groups, key=lambda bg: bg["start_col"])["end_col"],
        )
        if last_flow_groups
        else None
    )

    # Границы лекций для каждого потока
    flow_lecture_ranges = {}
    for bg in base_groups:
        flow_idx = bg["flow_idx"]
        if flow_idx not in flow_lecture_ranges:
            if flow_idx == max_flow_idx and last_group_no_lectures:
                groups_in_flow = sorted(
                    [g for g in base_groups if g["flow_idx"] == flow_idx],
                    key=lambda g: g["start_col"],
                )
                flow_lecture_ranges[flow_idx] = (
                    (bg["flow_start_col"], groups_in_flow[-2]["end_col"])
                    if len(groups_in_flow) >= 2
                    else None
                )
            else:
                flow_lecture_ranges[flow_idx] = (
                    bg["flow_start_col"],
                    bg["flow_end_col"],
                )

    header_row = course_range.max_row + 1
    day_col, time_col = course_range.min_col, course_range.max_col
    start_data_row = header_row + 1
    days_of_week = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]

    day_merged_ranges = [
        mr
        for mr in ws.merged_cells.ranges
        if mr.min_col == day_col
        and mr.min_row >= start_data_row
        and any(
            d in str(ws.cell(row=mr.min_row, column=mr.min_col).value or "").lower()
            for d in days_of_week
        )
    ]
    day_merged_ranges.sort(key=lambda r: r.min_row)

    days_schedule = []
    for i, day_mr in enumerate(day_merged_ranges):
        day_name = str(ws.cell(row=day_mr.min_row, column=day_mr.min_col).value).strip()
        day_min_row = day_mr.min_row
        day_index = next(
            (idx + 1 for idx, d in enumerate(days_of_week) if d in day_name.lower()),
            i + 1,
        )
        next_day_min_row = (
            day_merged_ranges[i + 1].min_row
            if i + 1 < len(day_merged_ranges)
            else ws.max_row + 1
        )

        day_time_slots = []
        for r in range(day_min_row, next_day_min_row):
            val = ws.cell(row=r, column=time_col).value
            if val is not None and str(val).strip() != "":
                day_time_slots.append({"label": str(val).strip(), "start_row": r})

        # Тот же принцип, что и для 1-2 курса — см. комментарий там.
        for j, slot in enumerate(day_time_slots):
            natural_end = (
                day_time_slots[j + 1]["start_row"] - 1
                if j + 1 < len(day_time_slots)
                else next_day_min_row - 1
            )
            slot["end_row"] = min(
                natural_end,
                slot["start_row"] + STANDARD_SLOT_HEIGHT + MAX_SLOT_ROW_SLACK - 1,
            )

        days_schedule.append(
            {"day_name": day_name, "day_index": day_index, "time_slots": day_time_slots}
        )

    all_parsed = []
    # Для 3-4 курса профилизации теперь считаются просто подгруппами
    # Поэтому мы перебираем базовые группы и передаем их на парсинг как обычно.
    for bg in base_groups:
        base_name, g_start_col, g_end_col = (
            bg["base_name"],
            bg["start_col"],
            bg["end_col"],
        )
        flow_idx = bg["flow_idx"]

        is_single = flow_counts[flow_idx] == 1

        is_lecture_participant = True
        lecture_range = flow_lecture_ranges.get(flow_idx)
        if (
            last_group_no_lectures
            and (g_start_col, g_end_col) == last_group_no_lectures
        ):
            is_lecture_participant, lecture_range = False, None
        if not lecture_range:
            is_lecture_participant = False

        for day in days_schedule:
            for slot in day["time_slots"]:
                rect = (slot["start_row"], g_start_col, slot["end_row"], g_end_col)
                entry_list = extract_class(
                    ws,
                    rect,
                    base_name,
                    course_number,
                    day["day_index"],
                    slot["label"],
                    lecture_coverage_start=lecture_range[0] if lecture_range else None,
                    lecture_coverage_end=lecture_range[1] if lecture_range else None,
                    is_lecture_participant=is_lecture_participant,
                    is_single_group_flow=is_single,
                )
                if entry_list:
                    all_parsed.extend(entry_list)

    return all_parsed


# ============================================================
# 8.  ИЗВЛЕЧЕНИЕ КЛАССА (ОБЩАЯ ДЛЯ ВСЕХ КУРСОВ)
# ============================================================


def extract_class(
    ws: Worksheet,
    rect: Tuple[int, int, int, int],
    group_name: str,
    course: int,
    day_of_week: int,
    time_label: str,
    lecture_coverage_start: Optional[int] = None,
    lecture_coverage_end: Optional[int] = None,
    is_lecture_participant: bool = True,
    is_single_group_flow: bool = False,
) -> List[Dict]:
    min_row, min_col, max_row, max_col = rect

    # Теперь мы не требуем, чтобы высота была ровно 4 строки, т.к. таймслот может быть 2 строки (например физкультура)
    if max_col - min_col + 1 != 2:
        return []

    # Нормализация высоты: стандартный блок пары — STANDARD_SLOT_HEIGHT (4) строк.
    # Если высота больше (граница на уровне дня даёт запас в MAX_SLOT_ROW_SLACK строк
    # на случай ошибки разметки таблицы или "утекающей" границы последнего слота дня),
    # решаем, какие именно 4 строки реальные:
    #   - первая строка непустая (для колонок ЭТОЙ группы) → лишняя строка снизу, берём первые 4;
    #   - первая строка пустая → лишняя строка сверху (ошибка в таблице), берём последние 4.
    # Слоты высотой <= STANDARD_SLOT_HEIGHT (например, 2-3 строки физкультуры) не трогаем.
    if max_row - min_row + 1 > STANDARD_SLOT_HEIGHT:
        first_row_values = [
            get_cell_value(ws, min_row, c) for c in range(min_col, max_col + 1)
        ]
        first_row_empty = not any(
            v is not None and str(v).strip() != "" for v in first_row_values
        )
        if first_row_empty:
            min_row = max_row - STANDARD_SLOT_HEIGHT + 1
        else:
            max_row = min_row + STANDARD_SLOT_HEIGHT - 1

    values_all = get_rect_values(ws, min_row, min_col, max_row, max_col)
    if not any(
        v is not None and str(v).strip() != "" for row in values_all for v in row
    ):
        return []

    start_time, end_time = parse_time_range(time_label)

    # 1. Физкультура и полные заливки (они не лекции)
    covering_merge = next(
        (
            mr
            for mr in ws.merged_cells.ranges
            if mr.min_row <= min_row
            and mr.max_row >= max_row
            and mr.min_col <= min_col
            and mr.max_col >= max_col
        ),
        None,
    )
    if covering_merge:
        title = ws.cell(row=covering_merge.min_row, column=covering_merge.min_col).value
        return [
            {
                "groupName": group_name,
                "course": course,
                "dayOfWeek": day_of_week,
                "startTime": start_time,
                "endTime": end_time,
                "isCommon": 1,
                "isLecture": 0,
                "classTitleA": str(title).strip() if title else "",
                "professorNameA": "",
                "classroomA": None,
                "classTitleB": None,
                "professorNameB": None,
                "classroomB": None,
                "comments": None,
            }
        ]

    # 2. Идеальная работа с лекциями.
    # ВАЖНО: раньше здесь была ещё и "elif is_lecture_participant" ветка со слабой
    # эвристикой (искала любой merge, выходящий за пределы ТЕКУЩЕЙ группы вправо).
    # Она использовалась курсом 1-2, который не передавал lecture_coverage_start/end.
    # Проблема: для ПОСЛЕДНЕЙ группы потока правая граница лекционного merge'а совпадает
    # с правой границей самой группы, условие "выходит за пределы" не срабатывало,
    # и лекция у последней группы потока разваливалась на "пару с 2 подгруппами"
    # (см. секцию 3 ниже), где имя преподавателя не отделялось от названия предмета.
    # Теперь и parse_course_1_2, и parse_course_3_4 всегда явно считают и передают
    # lecture_coverage_start/end на уровне потока, поэтому используется только
    # надёжная проверка "merge покрывает весь известный диапазон потока".
    lecture_merge = None
    if (
        is_lecture_participant
        and lecture_coverage_start is not None
        and lecture_coverage_end is not None
    ):
        lecture_merge = next(
            (
                mr
                for mr in ws.merged_cells.ranges
                if mr.min_row == min_row
                and mr.min_col <= min_col
                and mr.max_col >= max_col
                and mr.min_col <= lecture_coverage_start
                and mr.max_col >= lecture_coverage_end
            ),
            None,
        )

        # Если поток состоит всего из 1 группы, обычная 1-строчная пара совпадет с лекцией по ширине.
        # Отсекаем такие ложные "лекции", если высота объединения всего 1 строка.
        if is_single_group_flow and lecture_merge:
            if lecture_merge.max_row == min_row and lecture_merge.max_col == max_col:
                lecture_merge = None

    if lecture_merge:
        title = ws.cell(row=lecture_merge.min_row, column=lecture_merge.min_col).value

        # Если поток состоит из 1 группы, то это занятие физически не может быть классической лекцией
        final_is_lecture = 0 if is_single_group_flow else 1

        # Если лекция объединена по высоте полностью (или на 3-4 строки)
        if lecture_merge.max_row >= max_row or lecture_merge.max_row >= min_row + 2:
            return [
                {
                    "groupName": group_name,
                    "course": course,
                    "dayOfWeek": day_of_week,
                    "startTime": start_time,
                    "endTime": end_time,
                    "isCommon": 1,
                    "isLecture": final_is_lecture,
                    "classTitleA": str(title).strip() if title else "",
                    "professorNameA": "",
                    "classroomA": None,
                    "classTitleB": None,
                    "professorNameB": None,
                    "classroomB": None,
                    "comments": None,
                }
            ]

        # Стандартная лекция: безопасно проверяем, есть ли вообще 3 и 4 строки
        prof_val = (
            get_cell_value(ws, min_row + 2, lecture_merge.min_col)
            if min_row + 2 <= max_row
            else None
        )
        lecture_end_col = lecture_merge.max_col

        classroom_val = (
            get_cell_value(ws, min_row + 3, lecture_end_col - 1)
            if min_row + 3 <= max_row
            else None
        )
        if not classroom_val:
            classroom_val = (
                get_cell_value(ws, min_row + 3, lecture_end_col)
                if min_row + 3 <= max_row
                else None
            )
        # Раньше здесь было .replace(" ", ""), которое склеивало "130 ФМО" в "130ФМО".
        # Убираем только лишние/повторяющиеся пробелы по краям и внутри, сохраняя
        # значащий пробел между номером аудитории и названием корпуса.
        classroom = (
            re.sub(r"\s+", " ", str(classroom_val)).strip() if classroom_val else None
        )

        comments_values = []
        if min_row + 3 <= max_row:
            for c in range(lecture_merge.min_col, lecture_end_col - 1):
                merged = get_merged_range_for_cell(ws, min_row + 3, c)
                if merged and merged.min_col != c:
                    continue
                val = ws.cell(row=min_row + 3, column=c).value
                if val and str(val).strip():
                    comments_values.append(str(val).strip())
        comments = " ".join(comments_values) if comments_values else None

        return [
            {
                "groupName": group_name,
                "course": course,
                "dayOfWeek": day_of_week,
                "startTime": start_time,
                "endTime": end_time,
                "isCommon": 1,
                "isLecture": final_is_lecture,
                "classTitleA": str(title).strip() if title else "",
                "professorNameA": str(prof_val).strip() if prof_val else "",
                "classroomA": classroom,
                "classTitleB": None,
                "professorNameB": None,
                "classroomB": None,
                "comments": comments,
            }
        ]

    # 3. Обычная пара или пара с подгруппами (работает с динамической высотой)
    def get_row_safe(vals, idx):
        return vals[idx] if idx < len(vals) else [None, None]

    row0 = get_row_safe(values_all, 0)
    row1 = get_row_safe(values_all, 1)
    row2 = get_row_safe(values_all, 2)
    row3 = get_row_safe(values_all, 3)

    def clean_val(v):
        return str(v).strip() if v is not None and str(v).strip() != "" else None

    title_a, title_b = clean_val(row0[0]), clean_val(row0[1])
    prof_a = (
        " ".join(dict.fromkeys(filter(None, [clean_val(row1[0]), clean_val(row2[0])])))
        or None
    )
    prof_b = (
        " ".join(dict.fromkeys(filter(None, [clean_val(row1[1]), clean_val(row2[1])])))
        or None
    )
    classroom_a, classroom_b = clean_val(row3[0]), clean_val(row3[1])

    # УДАЛЕНИЕ ПРИЗРАЧНЫХ ПОДГРУПП:
    # Если названия предметов совпадают (или растянуты), но у одной стороны нет препода и аудитории,
    # мы полностью очищаем эту сторону (оставляя только реальную пару у нужной подгруппы).
    if title_a and title_b and title_a.lower() == title_b.lower():
        if not prof_a and not classroom_a and (prof_b or classroom_b):
            title_a = None
        if not prof_b and not classroom_b and (prof_a or classroom_a):
            title_b = None

    title_split = title_a != title_b
    prof_split = prof_a != prof_b
    classroom_split = classroom_a != classroom_b

    if not (title_split or prof_split or classroom_split):
        return [
            {
                "groupName": group_name,
                "course": course,
                "dayOfWeek": day_of_week,
                "startTime": start_time,
                "endTime": end_time,
                "isCommon": 1,
                "isLecture": 0,
                "classTitleA": title_a or title_b or "",
                "professorNameA": prof_a or prof_b,
                "classroomA": classroom_a or classroom_b,
                "classTitleB": None,
                "professorNameB": None,
                "classroomB": None,
                "comments": None,
            }
        ]

    return [
        {
            "groupName": group_name,
            "course": course,
            "dayOfWeek": day_of_week,
            "startTime": start_time,
            "endTime": end_time,
            "isCommon": 0,
            "isLecture": 0,
            "classTitleA": title_a or "",
            "professorNameA": prof_a,
            "classroomA": classroom_a,
            "classTitleB": title_b or "",
            "professorNameB": prof_b,
            "classroomB": classroom_b,
            "comments": None,
        }
    ]


# ============================================================
# 9.  ОТПРАВКА ДАННЫХ НА BACKEND
# ============================================================


def _group_fingerprint(records: List[Dict]) -> str:
    """Стабильный отпечаток содержимого группы, не зависящий от порядка записей."""
    canonical_records = sorted(
        records,
        key=lambda item: json.dumps(
            item, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        ),
    )
    canonical_json = json.dumps(
        canonical_records,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(canonical_json.encode("utf-8")).hexdigest()


def _post_import(url: str, headers: Dict[str, str], payload: Dict) -> Dict:
    last_error = None
    for attempt in range(1, 4):
        try:
            response = requests.post(url, json=payload, headers=headers, timeout=120)
            response.raise_for_status()
            return response.json()
        except (requests.exceptions.RequestException, ValueError) as error:
            last_error = error
            print(f"  ⚠️ Ошибка (попытка {attempt}/3): {error}")
            if attempt < 3:
                time.sleep(5)
    raise RuntimeError(f"Backend не принял данные: {last_error}")


def send_to_backend(parsed: List[Dict], course: int, allow_removals: bool = True):
    if not BACKEND_URL or not BACKEND_AUTH_TOKEN:
        return

    url = f"{BACKEND_URL.rstrip('/')}/api/schedule/import"
    headers = {"Authorization": f"Bearer {BACKEND_AUTH_TOKEN}"}
    by_group = {
        group_name: list(group_records)
        for group_name, group_records in groupby(
            sorted(parsed, key=lambda item: item["groupName"]),
            key=lambda item: item["groupName"],
        )
    }
    print(f"\nПроверка {len(by_group)} групп курса {course} на backend: {url}")

    state_result = _post_import(
        url, headers, {"mode": "state", "course": course}
    )
    old_groups = state_result.get("data", {}).get("groups", {})
    next_groups = {}
    changed_groups = []

    # Одна группа — одна небольшая порция и одна неизменяемая запись KV.
    # Сравнение выполняется здесь, чтобы не расходовать CPU Cloudflare Worker.
    for group_name, records in by_group.items():
        fingerprint = _group_fingerprint(records)
        old_entry = old_groups.get(group_name)
        if old_entry and old_entry.get("fingerprint") == fingerprint:
            next_groups[group_name] = old_entry
            continue

        upload_result = _post_import(
            url,
            headers,
            {
                "mode": "group",
                "course": course,
                "group": group_name,
                "fingerprint": fingerprint,
                "previousVersion": old_entry.get("version") if old_entry else None,
                "classes": records,
            },
        )
        next_groups[group_name] = upload_result["data"]["entry"]
        changed_groups.append(group_name)
        print(f"  ✅ Обновлена группа {group_name}")

    removed_groups = sorted(set(old_groups) - set(by_group))
    if not allow_removals:
        for group_name in removed_groups:
            next_groups[group_name] = old_groups[group_name]
        removed_groups = []
    if not changed_groups and not removed_groups:
        print(f"✅ Курс {course}: изменений нет, KV не обновлялся.")
        return

    result = _post_import(
        url,
        headers,
        {
            "mode": "finalize",
            "course": course,
            "importId": str(uuid.uuid4()),
            "groups": next_groups,
        },
    )
    if removed_groups:
        print(f"  🗑️ Удалены отсутствующие группы: {', '.join(removed_groups)}")
    print(
        f"✅ Курс {course} опубликован: изменено групп {len(changed_groups)}, "
        f"версия {result.get('version', 'unknown')}"
    )


# ============================================================
# 10. ОБЩАЯ ФУНКЦИЯ ПАРСИНГА
# ============================================================


def parse_excel_file(file_content: bytes, course_number: int) -> List[Dict]:
    if course_number in [1, 2]:
        return parse_course_1_2(file_content, course_number)
    elif course_number in [3, 4]:
        return parse_course_3_4(file_content, course_number)
    else:
        print(f"Курс {course_number} не поддерживается.")
        return []


# ============================================================
# 11. ОСНОВНАЯ ФУНКЦИЯ
# ============================================================


def main():
    print("Начало работы парсера...")
    service = get_drive_service()
    files = list_files_in_folder(service, DRIVE_FOLDER_ID)

    if not files:
        print("Нет файлов для обработки.")
        return

    all_parsed = []
    failed_files = []
    for file_meta in files:
        file_id, file_name = file_meta["id"], file_meta["name"]
        print(f"\nФайл: {file_name}")

        course_num = extract_course_number(file_name)
        if course_num is None:
            continue

        try:
            file_bytes = read_file_as_bytes(service, file_id)
            parsed = parse_excel_file(file_bytes, course_num)
            if parsed:
                all_parsed.extend(parsed)
        except Exception as e:
            print(f"  ❌ Ошибка при обработке файла '{file_name}': {e}")
            failed_files.append(file_name)
            continue

    if failed_files:
        print(
            f"\n⚠️ Не удалось обработать {len(failed_files)} файл(ов): {', '.join(failed_files)}"
        )

    if all_parsed:
        output_path = Path(__file__).parent / "parsed_schedule.json"
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(all_parsed, f, ensure_ascii=False, indent=4)
        print(f"\nВсе данные сохранены в {output_path}")

        if BACKEND_URL and BACKEND_AUTH_TOKEN:
            sorted_data = sorted(all_parsed, key=lambda x: x["course"])
            for course, group in groupby(sorted_data, key=lambda x: x["course"]):
                send_to_backend(list(group), course, allow_removals=not failed_files)
    print("\nПарсер завершил работу.")


if __name__ == "__main__":
    main()
